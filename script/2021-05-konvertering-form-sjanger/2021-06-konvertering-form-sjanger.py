from contextlib import contextmanager
from pathlib import Path
import pickle
import sys
from datetime import datetime
from typing import List, Optional, ContextManager, Tuple, Generator, Union

import pyodbc
import yaml
import attr
from dotenv import load_dotenv
from openpyxl import load_workbook
from structlog import get_logger

from seiso.console.helpers import Report, storage_path, ReportHeader
from seiso.services.nb import Vocabulary, Concept, Label
from seiso.services.promus import Promus
from seiso.services.promus.promus import MsSql

dry_run = True

load_dotenv()
log = get_logger()

vocabulary_code = 'ntsf'
vocabulary_name = 'Norsk tesaurus for sjanger og form'

script_dir = Path(__file__).parent

# $ monitor_nbvok
# $ cp ntsf.pickle script/2021-05-konvertering-form-sjanger/
pickle_file = script_dir.joinpath('%s.pickle' % vocabulary_code)

mapping_file = script_dir.joinpath('sjanger_form_4.xlsx')

with open(script_dir.joinpath('definitions.yml')) as fp:
    definitions = yaml.safe_load(fp)


def load_nb_dump() -> Vocabulary:
    try:
        with pickle_file.open('rb') as fp:
            return pickle.load(fp)
    except IOError as exc:
        return Vocabulary(marc_code=vocabulary_code, concepts={})


@contextmanager
def promus_conn() -> ContextManager[MsSql]:
    client = Promus()
    conn = client.connection()
    try:
        yield conn
    finally:
        conn.close()


promus = Promus()
global_promus_connection = promus.connection()


@contextmanager
def promus_cursor() -> ContextManager[pyodbc.Cursor]:
    with global_promus_connection.cursor() as cursor:
        yield cursor


@attr.s(auto_attribs=True)
class BibbiGenre(Concept):
    vocabulary: str = "bibbi"
    local_id: Optional[int] = None
    bibbi_id: Optional[int] = None


def load_bibbi() -> Generator[BibbiGenre, None, None]:
    query = """
        SELECT 
            TopicId AS local_id,
            authority.Bibsent_ID AS id, 
            Title AS label
        FROM AuthorityGenre AS authority
        WHERE ISNULL(GeoUnderTopic, '') = '' AND ISNULL(ReferenceNr, '') = '' AND NotInUse = 0
        ORDER BY authority.Bibsent_ID
    """
    with promus_cursor() as cursor:
        cursor.execute(query)
        for row in cursor:
            yield BibbiGenre(
                local_id=row[0],
                bibbi_id=row[1],
                label_nb=Label(row[2])
            )


def get_definition_sql(name: str) -> Optional[str]:
    def get_sql(x):
        for definition in definitions:
            if definition['name'] == x:
                return definition['sql']
        raise Exception('Not defined: ' + x)
    out = []
    for part in name.split(' '):
        if part == 'IKKE':
            if len(out) == 0:
                out.append('NOT')
            else:
                out.append('AND NOT')
        elif part == 'OG':
            out.append('AND')
        else:
            out.append(get_sql(part))
    return ' '.join(out)


def list_items(bibbi_concept: BibbiGenre, ekstra_kriterium: Optional[str]):
    query = """
        SELECT Item_ID, Title, Author, Varenr, Bibbinr
        FROM Item
        WHERE Item.Item_ID IN (
          SELECT Item.Item_ID
          FROM Item
          INNER JOIN ItemField AS f655 ON Item.Item_ID = f655.Item_ID
          WHERE f655.FieldCode = '655' AND f655.Authority_ID = ?
        )
    """
    if ekstra_kriterium:
        query += " AND (" + get_definition_sql(ekstra_kriterium) + ")"
    items = []
    print(query, [bibbi_concept.local_id])
    with promus_conn() as conn:
        for row in conn.select(query, [bibbi_concept.local_id], normalize=False):
            items.append(row)
    log.info('%d poster lest fra Promus' % len(items))
    return items

# 
# def get_local_genre_id(bibbi_id: int) -> int:
#     """Lookup local id from global id"""
#     with promus_cursor() as cursor:
#         cursor.execute("SELECT TopicID FROM AuthorityGenre WHERE Bibsent_ID = ?", [bibbi_id])
#         row = cursor.fetchone()
#         print(row)
#         print(row['TopicID'])
#         return row['TopicId']


def create_new_genre_authority(concept: Concept) -> BibbiGenre:
    new_bibbi_id, = update_query("EXEC sp_GetNextSequenceValue @Sequence_ID=?", (1,))
    now = datetime.now().isoformat(sep=' ', timespec='milliseconds')
    new_row = {
        'Title': concept.label_nb.value,
        'Title_N': concept.label_nn.value if concept.label_nn else None,
        'FieldCode': '655',
        'LastChanged': now,
        'Created': now,
        '_DisplayValue': concept.label_nb.value,
        'NotInUse': 0,
        'Bibsent_ID': new_bibbi_id,  # @TODO: Finne ut hvor denne genereres fra
        'Source': concept.vocabulary,
        'URI': concept.uri,
        'ConceptGroup': concept.group,
    }
    new_local_id = insert_query(
        'INSERT INTO AuthorityGenre (%s) VALUES (%s)' % (
            ','.join(list(new_row.keys())),
            ','.join(['?' for _ in new_row])
        ),
        tuple(new_row.values())
    )
    return BibbiGenre(**{**attr.asdict(concept, recurse=False), 'bibbi_id': new_bibbi_id, 'local_id': new_local_id})


def update_genre_authority(bibbi_concept: BibbiGenre, new_concept: Concept) -> BibbiGenre:
    update_query(
        'UPDATE AuthorityGenre SET Title=?, Title_N=?, Source=?, URI=?, ConceptGroup=? WHERE TopicID=?',
        (
            new_concept.label_nb.value,
            new_concept.label_nn.value if new_concept.label_nn else None,
            new_concept.vocabulary,
            new_concept.uri,
            new_concept.group,
            bibbi_concept.local_id
        )
    )
    return BibbiGenre(**{**attr.asdict(bibbi_concept, recurse=False), **attr.asdict(new_concept, recurse=False)})


def update_items_query(items: List, old_genre: int, new_genre: int):
    query = """
            UPDATE ItemField
            SET Authority_ID = %d
            WHERE f655.FieldCode = '655'
            AND f655.Authority_ID = %d
            AND f655.Item_ID IN (%s);
        """ % (new_genre, old_genre, ', '.join(['%d' % x for x in items]))
    with open('items_query.sql', 'a') as fp:
        fp.write(query)


def find_nb_concept(nb_vok: Vocabulary, term: str) -> Optional[Concept]:
    if not term:
        return
    for nb_uri, nb_concept in nb_vok.concepts.items():
        if nb_concept.label_nb.value == term:
            return nb_concept


def find_bibbi_concept(bibbi_vok, term: str) -> Optional[BibbiGenre]:
    if not term:
        return
    matches = [c for c in bibbi_vok if c.label_nb.value == term]
    if len(matches) > 1:
        print("FEIL: Mer enn ett Bibbi-begrep med termen: %s" % term)
    elif len(matches) == 1:
        return matches[0]


def update_query(query: str, params: tuple) -> Optional[tuple]:
    with promus_cursor() as cursor:
        print(query, params)
        if dry_run:
            return -1,
        else:
            print("NOT DRY RUN")
            cursor.execute(query, params)
            return cursor.fetchone()


def insert_query(query: str, params: tuple):
    with promus_cursor() as cursor:
        print(query, params)
        if not dry_run:
            print("NOT DRY RUN")
            cursor.execute(query, params)
            cursor.execute("SELECT @@IDENTITY")
            new_id = cursor.fetchone()[0]
            return new_id
        return -1


def main():
    nbvok = load_nb_dump()
    log.info('Leste %d begreper fra ntsf' % len(nbvok.concepts))
    bibbi = list(load_bibbi())
    log.info('Leste %d begreper fra bibbi' % len(bibbi))
    mapping_data = load_workbook(str(mapping_file)).active

    def get_cell_value(row, col):
        cell = mapping_data.cell(row=row, column=col)
        return cell.value or '' if cell else ''

    # (1) VALIDATE

    bibbi_term_and_filter = set()
    invalid_rows = set()
    for rowno in range(2, mapping_data.max_row):
        bibbi_term = get_cell_value(rowno, 1)
        if bibbi_term:
            bibbi_concept = find_bibbi_concept(bibbi, bibbi_term)
            ekstra_kriterium = get_cell_value(rowno, 2)
            # Sjekk at term + kriterium er unik
            idx = bibbi_term + ekstra_kriterium
            if idx in bibbi_term_and_filter:
                print("FEIL: Bibbi-term + kriterium må være unik:", bibbi_term, ekstra_kriterium)
                sys.exit(1)
            bibbi_term_and_filter.add(idx)
            if not bibbi_concept:
                print("Ugyldig bibbi-term: %s" % bibbi_term)
                invalid_rows.add(rowno)

        new_term = get_cell_value(rowno, 3)
        new_vocab = get_cell_value(rowno, 5)

        if new_vocab == 'ntsf':
            ntsf_concept = find_nb_concept(nbvok, new_term)
            if not ntsf_concept:
                print('FEIL: Ugyldig NTSF-term: "%s"' % new_term)
                invalid_rows.add(rowno)

    # (2) PROCESS

    report = Report()
    bibbi_terms_processed = set()
    for rowno in range(2, mapping_data.max_row):

        if rowno in invalid_rows:
            print("SKIPPING ROW", rowno)
            continue

        bibbi_term = get_cell_value(rowno, 1)
        bibbi_id = None
        ekstra_kriterium = get_cell_value(rowno, 2)
        new_term = get_cell_value(rowno, 3)
        new_term_nn = get_cell_value(rowno, 4)
        new_vocab = get_cell_value(rowno, 5)
        new_group = get_cell_value(rowno, 6)

        bibbi_concept = find_bibbi_concept(bibbi, bibbi_term)
        if bibbi_concept:
            bibbi_id = bibbi_concept.bibbi_id
            items = list_items(bibbi_concept, ekstra_kriterium)
        else:
            items = []

        new_authority = None
        row_status = ""

        if not new_term:
            continue

        if new_vocab == 'ntsf':
            ntsf_concept = find_nb_concept(nbvok, new_term)
            if bibbi_term:
                # Case 1: Bibbi-term mappet til NTSF-term

                if bibbi_term not in bibbi_terms_processed:
                    # I tilfeller der en Bibbi-term er mappet til flere NTSF-termer,
                    # er det den første NTSF-termen som skal brukes.
                    row_status = "Oppdatert"
                    new_authority = update_genre_authority(bibbi_concept, ntsf_concept)
                    bibbi_terms_processed.add(bibbi_term)
                else:
                    row_status = "Opprettet"

                    # CREATE NEW AUTHORITY
                    new_authority = create_new_genre_authority(ntsf_concept)
                    bibbi_id = new_authority.bibbi_id

                    # THEN REPLACE
                    update_items_query(
                        [item['Item_ID'] for item in items],
                        bibbi_concept.local_id,
                        new_authority.local_id
                    )
            else:
                row_status = "Opprettet"
                new_authority = create_new_genre_authority(ntsf_concept)
                bibbi_id = new_authority.bibbi_id

        elif new_vocab == 'bibbi':
            if bibbi_term and bibbi_term not in bibbi_terms_processed:
                row_status = "Oppdatert"
                new_authority = update_genre_authority(bibbi_concept, BibbiGenre(
                    label_nb=Label(new_term),
                    label_nn=Label(new_term_nn) if new_term_nn != '' else None,
                    group=1 if new_group == 'film' else '2',
                ))
                bibbi_terms_processed.add(bibbi_term)
            else:
                row_status = "Opprettet"
                new_authority = create_new_genre_authority(BibbiGenre(
                    label_nb=Label(new_term),
                    label_nn=Label(new_term_nn) if new_term_nn != '' else None,
                    group=1,
                ))
                bibbi_id = new_authority.bibbi_id

        if row_status:
            # Report row
            row = [
                row_status,
                str(bibbi_id) if bibbi_id is not None else "",
                bibbi_term or "",
            ]
            if new_authority:
                print(new_authority)
                row += [
                    new_authority.label_nb.value,
                    new_authority.vocabulary or "",
                    new_authority.uri or "",
                ]
            else:
                row += ["", "", ""]

            row += ['%d' % len(items)]

            print(row)
            report.add_row(row)

    reports_path = storage_path('reports')

    report.save_excel(reports_path.joinpath('nbvok-bibbi-sjanger.xlsx'), headers=[
        ReportHeader('', 'Handling', width=20),
        ReportHeader('', 'Bibbi-ID', width=40),
        ReportHeader('', 'Tidligere term', width=40),
        ReportHeader('', 'Ny term', width=40),
        ReportHeader('', 'Vokabular', width=40),
        ReportHeader('', 'URI', width=40),
        ReportHeader('', 'Antall poster', width=40),
    ])


if __name__ == '__main__':
    main()
