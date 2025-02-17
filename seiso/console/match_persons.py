from __future__ import annotations
import argparse
from dataclasses import asdict
from typing import Union
import logging
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from requests import Session
from cachecontrol import CacheControlAdapter
from cachecontrol.heuristics import ExpiresAfter
from ..services.alma import get_alma_candidates
from ..common.interfaces import NorafPerson, ViafPerson, Strategy, Match, NoMatch
from ..matcher.matchers import isbn_matcher, title_matcher
from ..services.viaf import get_viaf_candidates
from ..services.promus.authorities import Promus, BibbiPersons, BibbiPerson, QueryFilter
from ..common.logging import setup_logging

logger = setup_logging()


def match_person(bibbi_person: BibbiPerson) -> Union[Match, NoMatch]:
    strategies = [
        Strategy(
            name='isbn',
            provider=get_alma_candidates,
            query='alma.isbn="{isbn}"',
            matcher=isbn_matcher,
        ),
        Strategy(
            name='creator+title',
            provider=get_alma_candidates,
            query='alma.creator="{creator}" AND alma.title="{title}"',
            matcher=title_matcher,
        ),
        Strategy(
            name='creator+fuzzy title',
            provider=get_alma_candidates,
            query='alma.creator="{creator}"',
            matcher=title_matcher,
        ),
        Strategy(
            name='viaf:creator+title',
            provider=get_viaf_candidates,
            query='local.personalNames="{creator}"',
            matcher=title_matcher,
        ),
    ]

    viaf_match = None

    session = Session()
    session.mount('https://', CacheControlAdapter(
        heuristic=ExpiresAfter(days=1),
        max_retries=10
    ))

    for strategy in strategies:
        for bibbi_item in bibbi_person.items:
            logger.info(' -> {%s} Check item: %s %s', strategy.name, bibbi_item.isbn, ' || '.join(bibbi_item.titles))

            query = strategy.query.format(**{
                'creator': bibbi_person.name.strip(),
                'title': bibbi_item.titles[0].replace('"', '').strip(),
                **asdict(bibbi_item)
            })

            candidates = strategy.provider(query, session)

            for candidate in candidates:
                if match := strategy.matcher(bibbi_person, bibbi_item, candidate, strategy):
                    if isinstance(match.target, NorafPerson):
                        return match
                    elif isinstance(match.target, ViafPerson) and viaf_match is None:
                        # If we find a VIAF-only match, we keep it, but we will continue to check if we can
                        # find one with a Noraf link.
                        viaf_match = match

    if viaf_match is not None:
        return viaf_match

    return NoMatch()


def match_persons(persons: BibbiPersons):
    wb = Workbook()
    ws = wb.active

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 30
    ws.column_dimensions['K'].width = 30
    ws.column_dimensions['L'].width = 40
    ws.column_dimensions['M'].width = 40

    headers = [
        [
            'Bibbi-autoritet', '', '', '', '',
            'Noraf-autoritetskandidat', '', '',
            'Noraf-match basert på', '', '', '',
            'Viaf-autoritetskandidat', ''
        ],
        [
            'ID', 'Navn', 'Datoer', 'Antall poster', 'Nyeste post godkjent',
            'ID', 'Navn', 'Datoer', 'Strategi', 'ISBN/Tittel', 'Navn', 'Fødselsår',
            'ID', 'Tittel'
        ],
    ]

    header_font = Font(bold=True)

    for n1, v1 in enumerate(headers):
        for n2, v2 in enumerate(v1):
            cell = ws.cell(n1 + 1, n2 + 1, v2)
            cell.font = header_font
            cell.fill = PatternFill('solid', fgColor='FFFFEE')

    row = 3
    person_no = 0
    person_t = len(persons)
    for bibbi_id, bibbi_person in persons.items():
        person_no += 1
        logger.info('[%d/%d] %s %s' % (person_no, person_t, bibbi_id, bibbi_person))

        match = match_person(bibbi_person)
        if isinstance(match, NoMatch):
            logger.info(' => No match')
        elif isinstance(match, ViafPerson):
            logger.info(' => VIAF only match')
        else:
            logger.info(' => {%s} %s', match.strategy, match.title_similarity)
            # '\n  [strategy:%s] %s "%s"' % (strategy.name, bibbi_item.isbn, '" || "'.join(bibbi_item.titles)))

        ws.cell(row=row, column=1, value=bibbi_id)
        ws.cell(row=row, column=2, value=bibbi_person.name)
        ws.cell(row=row, column=3, value=bibbi_person.dates or '')
        ws.cell(row=row, column=4, value=len(bibbi_person.items))
        ws.cell(row=row, column=5, value=bibbi_person.newest_approved)

        ws.cell(row=row, column=6, value=match.strategy)

        if isinstance(match.target, NorafPerson):
            ws.cell(row=row, column=7, value=match.target.id)
            ws.cell(row=row, column=8, value=match.target.name)
            ws.cell(row=row, column=9, value=match.target.dates or '')
            ws.cell(row=row, column=10, value=match.title_similarity)
            ws.cell(row=row, column=11, value=match.name_similarity)
            ws.cell(row=row, column=12, value=match.date_similarity)

        if isinstance(match.target, ViafPerson):
            ws.cell(row=row, column=13, value='=HYPERLINK("https://viaf.org/viaf/%s")' % match.target.id)
            ws.cell(row=row, column=14, value=match.title_similarity)

        row += 1

        outfile = 'bibbi-persons-match-alma.xlsx'
        wb.save('results/%s' % outfile)
        logger.debug('Wrote %s', outfile)


def main():
    """
    Script for matching Bibbi persons to Noraf using a combination of Alma and VIAF APIs.
    The results are written to an Excel file.
    """
    load_dotenv()

    parser = argparse.ArgumentParser(description='Match Bibbi persons to Noraf using Alma + VIAF')
    parser.add_argument('--debug', action='store_true')
    args = parser.parse_args()

    if args.debug:
        logger.setLevel(logging.DEBUG)
    else:
        logger.setLevel(logging.INFO)

    logger.info('Starting')

    promus = Promus()

    # Hent alle personer fra Bibbi
    bibbi_persons = promus.authorities.person.list_records(
        [
            QueryFilter("ReferenceNr IS NULL"),
            QueryFilter("Felles_ID = Bibsent_ID"),
            QueryFilter("NB_ID IS NULL"),
        ]
    )
    logger.info('%d persons read from Promus', len(bibbi_persons))

    # Velg ut de som har minst en utgivelse i 2019 eller 2020, men spar på alle utgivelsene til disse personene,
    # så vi kan bruke dem til matching.
    bibbi_persons_filtered = {}
    for bibbi_id, bibbi_person in bibbi_persons.items():
        if bibbi_person.newest_approved is None:
            continue
        newest_year = int(bibbi_person.newest_approved.strftime('%Y'))
        if 2019 <= newest_year <= 2020:
            bibbi_persons_filtered[bibbi_id] = bibbi_person
    logger.info('%d persons have items published within requested date range', len(bibbi_persons_filtered))

    match_persons(bibbi_persons_filtered)
