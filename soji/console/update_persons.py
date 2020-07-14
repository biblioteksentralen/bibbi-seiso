import argparse
import logging
import time
import os
from dataclasses import dataclass
from pathlib import Path
from typing import Union

import questionary
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from soji.common.bare import Bare
from soji.common.bare_record import BibsysJsonMarcField

from soji.common.interfaces import BibbiPerson, BarePersonRecord
from soji.common.promus import Promus
from soji.common.logging import logger


# Because of https://youtrack.jetbrains.com/issue/PY-30747
PathType = Union[Path, os.PathLike]


@dataclass
class BibbiBareMatch():
    bibbi_id: str
    bare_id: str


nationality_map = {
    'n': 'no',
}


def update_person(promus: Promus, bare: Bare, bibbi_person: BibbiPerson, match: BibbiBareMatch, dryrun: bool,
                  backup_path: Path):

    bare_json_rec = bare.get(match.bare_id)
    bare_person: BarePersonRecord = bare_json_rec.simple_record()

    logger.debug('Sjekker: Bibbi: "%s" <=> BARE: "%s"', bibbi_person, bare_person)

    # --------------------------------------------------------------------------
    # Update Bibbi

    if bibbi_person.dates is None and bare_person.dates is not None:
        logger.warning('BIBBI %s: [TODO] Legger til dato fra BARE: %s', bibbi_person.id, bare_person.dates)
        # TODO

    logger.info('BIBBI %s: Setter BARE-ID = "%s"', bibbi_person.id, bare_person.id)
    if not dryrun:
        promus.persons.update(bibbi_person,
                              NB_ID=int(bare_person.id),
                              Handle_ID=bare_json_rec.identifiers('handle')[0].split('/', 3)[-1],
                              Origin=bare_json_rec.origin,
                              Kat=bare_json_rec.status,
                              NB_PersonNation=bare_json_rec.nationality
                              )

    # --------------------------------------------------------------------------
    # Update BARE

    if len(bare_json_rec.identifiers('bibbi')):
        bare_current_bibbi = bare_json_rec.identifiers('bibbi')[0]
        if bare_current_bibbi == bibbi_person.id:
            logger.debug('BARE %s: Allerede mappet til riktig Bibbi-post', bare_person.id)
        else:
            logger.warning('BARE %s: Mappet til annen Bibbi-post: %s', bare_person.id, bare_current_bibbi)
    else:

        # Make a backup first
        if not dryrun:
            backup_file_path: PathType = backup_path.joinpath('%s_before.json' % bare_person.id)
            if not backup_file_path.exists():
                with open(backup_file_path, 'w', encoding='utf-8') as fp:
                    fp.write(bare_json_rec.as_json())

        logger.info('BARE %s: Legger til Bibbi-ID = "%s"', bare_person.id, bibbi_person.id)
        bare_json_rec.set_identifiers('bibbi_id', [bibbi_person.id])

        # ------------------------------------------------------------------------------------------------------------
        # 100$d Datoer

        if bibbi_person.dates is not None and bare_person.dates is None:
            bare_json_rec.first('100').set('d', bibbi_person.dates)
            logger.info('BARE %s: Setter 100$d = "%s"', bare_person.id, bibbi_person.dates)  # "%s" <=> "%s"' % (bibbi_person, bare_rec_simple))
        elif bibbi_person.dates is not None and bare_person.dates is not None:
            if bibbi_person.dates != bare_person.dates:
                logger.warning('Ulike datoer: "%s" (BARE:%s) != "%s" (BIBBI:%s)',
                               bare_person.dates,
                               bare_person.id,
                               bibbi_person.dates,
                               bibbi_person.id)

        # ------------------------------------------------------------------------------------------------------------
        # 386 Nasjonalitet

        if bibbi_person.nationality is not None and bare_person.nationality is None:
            bare_json_rec.add(BibsysJsonMarcField.construct('386', ' ', ' ', [
                {
                    "subcode": "a",
                    "value": bibbi_person.nationality,
                },
                {
                    "subcode": "m",
                    "value": "Nasjonalitet/regional gruppe",
                },
                {
                    "subcode": "2",
                    "value": "bs-nasj",
                },
            ]))
            logger.info('BARE %s: Setter 386$a = "%s"', bare_person.id,
                        bibbi_person.nationality)  # "%s" <=> "%s"' % (bibbi_person, bare_rec_simple))

        # ------------------------------------------------------------------------------------------------------------
        # 086 $c Landskoder

        if len(bibbi_person.country_codes) != 0 and len(bare_person.country_codes) == 0:
            bare_json_rec.add(BibsysJsonMarcField.construct('043', ' ', ' ', [
                {
                    "subcode": "c",
                    "value": countryCode,
                }
                for countryCode in bibbi_person.country_codes
            ]))
            logger.info('[BARE:%s] Setter 043$c = %s',
                        bare_person.id, ', '.join(bibbi_person.country_codes))

        # ------------------------------------------------------------------------------------------------------------
        # Write result to file

        if not dryrun:
            bare.put(bare_json_rec)
            file_path: PathType = backup_path.joinpath('%s_after.json' % bare_person.id)
            bare_json_rec = bare.get(match.bare_id)
            with open(file_path, 'w', encoding='utf-8') as fp:
                fp.write(bare_json_rec.as_json())

            time.sleep(5)


def main():
    """
    Scriptet oppdaterer personposter i Bibbi (via SQL) og BARE (via REST-API) basert på inputt
    fra Excel-filen bare_forslag_fra_isbn_tittel_match.xlsx
    """
    parser = argparse.ArgumentParser(description='Update persons using data from Excel sheet')
    parser.add_argument(
        'infile',
        nargs='?',
        default=r'/mnt/c/Users/dmheg/Bibliotekenes IT-Senter/Bibliotekfaglig avdeling - General/' +
                r'Promusdugnad/bare_forslag_fra_isbn_tittel_match.xlsx',
        type=argparse.FileType('rb')
    )
    parser.add_argument('-n', '--dry', action='store_true', help='Dry run: Show the update operations the script ' +
                                                                 'would perform, without actually performing them.')
    parser.add_argument('-v', '--verbose', action='store_true', help='More verbose output.')
    args = parser.parse_args()

    if args.verbose:
        logger.setLevel(logging.DEBUG)
    else:
        logger.setLevel(logging.INFO)

    load_dotenv()
    promus = Promus()
    bare = Bare(os.getenv('BARE_KEY'))

    wb = load_workbook(args.infile)

    sheet: Worksheet = wb['Resultater']

    logger.info('Read %d rows, %d columns' % (sheet.max_row, sheet.max_column))
    marked_ok = []
    marked_not_ok = []
    for row in sheet.iter_rows(3):
        if row[0].value is None:
            continue

        status = str(row[0].value)

        if status.lower() == 'ok':
            marked_ok.append(BibbiBareMatch(
                bibbi_id=str(row[1].value),
                bare_id=str(row[6].value)
            ))
        else:
            logger.debug('Annen status: %s', status)
            marked_not_ok.append(BibbiBareMatch(
                bibbi_id=str(row[1].value),
                bare_id=str(row[6].value)
            ))

    logger.info('Records marked OK: %d, not OK: %d', len(marked_ok), len(marked_not_ok))

    backup_path = Path('./', 'backup')
    logger.info('Backup path: %s', backup_path.absolute())
    backup_path.mkdir(exist_ok=True)

    if args.dry:
        logger.info('Running in dry-run mode. No actual changes will be carried out.')
        time.sleep(2)
    else:
        if not questionary.confirm('Scriptet vil gjøre endringer i poster. Sikker på at du vil fortsette?', default=False).ask():
            return

    n = 0
    for match in marked_ok:
        bibbi_person = promus.persons.get(match.bibbi_id)
        if bibbi_person is None:
            logger.warning('Person no longer exists: %s', match.bibbi_id)
        elif bibbi_person.bare_id is not None:
            logger.warning('Person already matched to %s. Our suggestion: %s', bibbi_person.bare_id, match.bare_id)
        else:
            update_person(promus, bare, bibbi_person, match, args.dry, backup_path)
            n += 1
            if n > 3:
                return
