import requests
import pydash
import pyodbc
import os
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


class MsSql:

    def __init__(self, **db_settings):
        if os.name == 'posix':
            conn_string = ';'.join([
                'DRIVER={FreeTDS}',
                'Server=%(server)s',
                'Database=%(database)s',
                'UID=BIBSENT\\dmheg',
                'PWD=%(password)s',
                'TDS_Version=8.0',
                'Port=1433',
            ]) % db_settings
        else:
            conn_string = ';'.join([
                'Driver={SQL Server}',
                'Server=%(server)s',
                'Database=%(database)s',
                'Trusted_Connection=yes',
            ]) % db_settings
        self.conn: pyodbc.Connection = pyodbc.connect(conn_string)

    def cursor(self) -> pyodbc.Cursor:
        return self.conn.cursor()


def fetch_persons_from_promus() -> dict:
    return []

    db = MsSql(server=os.getenv('DB_SERVER'), database=os.getenv('DB_DB'),
               user=os.getenv('DB_USER'), password=os.getenv('DB_PASSWORD'))

    query = """
    SELECT person.Bibsent_ID, person.PersonName, person.PersonYear, item.Varenr, item.Title
    FROM AuthorityPerson AS person

    LEFT JOIN ItemField AS field ON field.Authority_ID = person.PersonId AND field.FieldCode IN ('100', '600')
    LEFT JOIN Item AS item ON field.Item_ID = item.Item_ID

    WHERE person.ReferenceNr IS NULL AND person.Approved = '1' AND person.Felles_ID = person.Bibsent_ID AND person.NB_ID IS NULL
    AND item.Varenr IS NOT NULL
    """

    persons: dict = {}
    with db.cursor() as cursor:
        cursor.execute(query)
        for row in cursor:
            bibbi_id = row[0]
            if bibbi_id not in persons:
                if dates == '':
                    dates = None
                persons[bibbi_id] = {
                    'name': row[1],
                    'dates': dates,
                    'titles': [],
                }
            persons[bibbi_id]['titles'].append(row[3])
    return persons


def search_alma_for_person(bibbi_person: dict) -> dict:

    search_tpl = 'https://ub-lsm.uio.no/alma/search?query=alma.all_for_ui%3D%22{}%22&expand_items=false&nz=true'

    matches = {}
    for isbn in bibbi_person['titles']:
        url = search_tpl.format(isbn)
        res = requests.get(url).json()
        for bare_person in pydash.get(res, 'results.0.creators', []):
            bare_id = bare_person['id'].replace('(NO-TrBIB)', '')
            if bare_person['name'] == bibbi_person['name'] and bibbi_person['dates'] == bare_person.get('dates'):
                matches[bare_id] = ('exact', isbn, bare_person)
                break
            elif bare_person['name'] == bibbi_person['name'] and bibbi_person['dates'] != bare_person.get('dates'):
                matches[bare_id] = ('fuzzy', isbn, bare_person)
                break
    if len(matches) == 0:
        matches['-'] = ('no_match', '', {'name': '', 'dates': ''})

    return matches


def match_persons(persons: list):
    wb = Workbook()
    ws = wb.active

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 30

    headers = [
        ['Bibbi-autoritet', '', '', 'BARE-autoritet', '', '','Match',''],
        ['ID', 'Navn', 'Datoer',  'ID', 'Navn', 'Datoer', 'Sikkerhet', 'Eksempel-ISBN'],
    ]

    header_font = Font(bold=True)

    for n1, v1 in enumerate(headers):
        for n2, v2 in enumerate(v1):
            cell = ws.cell(n1 + 1, n2 + 1, v2)
            cell.font = header_font
            cell.fill = PatternFill("solid", fgColor="FFFFEE")


    row = 3
    for bibbi_id, bibbi_person in persons.items():
        res = search_alma_for_person(bibbi_person)
        for k,v in res.items():

            ws.cell(row=row, column=1, value=bibbi_id)
            ws.cell(row=row, column=2, value=bibbi_person['name'])
            ws.cell(row=row, column=3, value=bibbi_person['dates'])
            ws.cell(row=row, column=4, value=k)
            ws.cell(row=row, column=4, value=v[2]['name'])
            ws.cell(row=row, column=4, value=v[2]['dates'])
            ws.cell(row=row, column=4, value=v[0])
            ws.cell(row=row, column=4, value=v[1])

            row += 1

        wb.save('bibbi-persons-match-alma.xlsx')



load_dotenv()
bibbi_persons = fetch_persons_from_promus()

# bibbi_persons = {
#     'TODO': {
#         'name': 'Thoene, Bodie',
#         'dates': '1951-',
#         'titles': [
#             '8273410706'
#         ]
#     },
#     'TODO1': {
#         'name': 'Lange, Sven',
#         'dates':  '1957-',
#         'titles': [
#             '9788242110497'
#         ]
#     }
# }

match_persons(bibbi_persons)


