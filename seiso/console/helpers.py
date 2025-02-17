import json
import logging
import os
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill
from tqdm import tqdm

from seiso.common.logging import setup_logging

log = setup_logging(level=logging.INFO)


@dataclass
class ReportHeader(object):
    line1: str
    line2: str = ''
    width: int = 50


def storage_path(dirname: Optional[str] = None, create: bool = True) -> Path:
    STORAGE_PATH = os.getenv("STORAGE_PATH")
    if STORAGE_PATH is None:
        raise ValueError("STORAGE_PATH environment variable not set")
    path = Path(STORAGE_PATH).expanduser()
    if dirname is not None:
        path = path.joinpath(dirname)
    if create:
        path.mkdir(exist_ok=True, parents=True)
    return path


def log_path(filename: str) -> Path:
    return Path(os.getenv('LOG_PATH', '')).expanduser().joinpath(filename)


class Report:

    def __init__(self):
        self.data = []

    def add_row(self, columns):
        for col in columns:
            if not isinstance(col, str):
                log.warning('Warning: Column not of type str: %s', str(columns))
        self.data.append(columns)

    def save_json(self, filename: Path):
        with filename.open('w', encoding='utf-8') as fp:
            json.dump(self.data, fp, indent=2, ensure_ascii=False)
        log.info('Saved %d data rows to %s', len(self.data), filename)

    def load_json(self, filename: Path):
        with filename.open('r', encoding='utf-8') as fp:
            self.data = json.load(fp)

    def save_excel(self, filename: Path, headers=None):
        log.info('Writing Excel file')
        headers = headers or []
        wb = Workbook(write_only=True)
        ws = wb.create_sheet()

        for n, header in enumerate(headers):
            ws.column_dimensions[chr(65 + n)].width = header.width

        ws.freeze_panes = 'A3'

        header_font = Font(bold=True)
        header_fill = PatternFill('solid', fgColor='FFFFEE')
        link_font = Font(color='0000FF')

        row1 = []
        row2 = []
        for header in headers:
            cell1 = WriteOnlyCell(ws, header.line1)
            cell1.font = header_font
            cell1.fill = header_fill
            row1.append(cell1)
            cell2 = WriteOnlyCell(ws, header.line2)
            cell2.font = header_font
            cell2.fill = header_fill
            row2.append(cell2)
        ws.append(row1)
        ws.append(row2)

        for row_no, values in enumerate(tqdm(self.data)):
            row = []
            for value in values:
                link = None
                if value is None:
                    value = ''
                elif value.startswith('{BIBBI}'):
                    value = value[7:]
                    link = 'https://id.bibbi.dev/bibbi/' + value
                elif value.startswith('{NORAF}'):
                    value = value[7:]
                    link = 'https://bsaut.toolforge.org/show/' + value

                cell = WriteOnlyCell(ws, value)

                if link is not None:
                    cell.hyperlink = link
                    cell.font = link_font

                row.append(cell)
            ws.append(row)

        last_row_no = len(self.data) + 2
        last_col_chr = chr(64 + len(headers))
        ws.auto_filter.ref = "A2:%s%s" % (last_col_chr, last_row_no)

        log.info('Saving')
        wb.save(filename)
        log.info('Wrote %d data rows to %s', len(self.data), filename)
