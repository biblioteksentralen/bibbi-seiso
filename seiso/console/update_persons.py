import argparse
import logging
import re
import time
import os
from dataclasses import dataclass
from pathlib import Path
from typing import Union, Optional

import questionary
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from seiso.services.noraf import Noraf
from seiso.common.noraf_record import NorafJsonMarcField

from seiso.common.interfaces import BibbiPerson, NorafPersonRecord
from seiso.services.promus import Promus
from seiso.common.logging import setup_logging

logger = setup_logging()

# Because of https://youtrack.jetbrains.com/issue/PY-30747
PathType = Union[Path, os.PathLike]


@dataclass
class BibbiNorafMatch():
    bibbi_id: str
    noraf_id: str


def update_person(promus: Promus,
                  noraf: Noraf,
                  match: BibbiNorafMatch,
                  dry_run: bool,
                  backup_path: Optional[Path] = None):

    bibbi_person = promus.persons.get(match.bibbi_id)

    if bibbi_person is None:
        logger.info('Person no longer exists in Bibbi: %s', match.bibbi_id)
        return False

    if bibbi_person.noraf_id is not None:
        if bibbi_person.noraf_id != match.noraf_id:
            logger.warning('Person already matched to %s. Our suggestion: %s', bibbi_person.noraf_id, match.noraf_id)
        return False

    noraf_json_rec = noraf.get(match.noraf_id)
    noraf_person: NorafPersonRecord = noraf_json_rec.simple_record()

    if len(noraf_json_rec.identifiers('bibbi')):
        noraf_current_bibbi = noraf_json_rec.identifiers('bibbi')[0]
        if noraf_current_bibbi == bibbi_person.id:
            logger.debug('[Noraf:%s] Allerede mappet til riktig Bibbi-post', noraf_person.id)
        else:
            logger.warning('[Noraf:%s] Mappet til en annen Bibbi-post: %s', noraf_person.id, noraf_current_bibbi)
        return False

    logger.info('--- Bibbi: "%s" <> Noraf: "%s" ---', bibbi_person, noraf_person)

    # ==========================================================================
    # Update Bibbi
    # ==========================================================================

    if bibbi_person.dates is None and noraf_person.dates is not None:
        logger.warning('[Bibbi:%s] [TODO] Legger til dato fra Noraf: %s', bibbi_person.id, noraf_person.dates)
        # TODO

    logger.info('[Bibbi:%s] Setter Noraf-ID = "%s"', bibbi_person.id, noraf_person.id)
    promus.persons.link_to_noraf(bibbi_person, noraf_json_rec, dry_run, reason='update_persons.py')

    if bibbi_person.gender is None and noraf_person.gender is not None:
        logger.info('[Bibbi:%s] Setter gender = "%s"', bibbi_person.id, noraf_person.gender)
        promus.persons.update(bibbi_person,
                              dry_run,
                              Gender=noraf_person.gender)

    # ==========================================================================
    # Update Noraf
    # ==========================================================================

    # Make a backup first
    if not dry_run:
        if backup_path is not None:
            backup_file_path: PathType = backup_path.joinpath('%s_before.json' % noraf_person.id)
            if not backup_file_path.exists():
                with open(backup_file_path, 'w', encoding='utf-8') as fp:
                    fp.write(noraf_json_rec.as_json())

    logger.info('[Noraf:%s] Setter Bibbi-ID = "%s"', noraf_person.id, bibbi_person.id)
    noraf_json_rec.set_identifiers('bibbi', [bibbi_person.id])

    # ------------------------------------------------------------------------------------------------------------
    # 100$d Datoer

    if bibbi_person.dates is not None and noraf_person.dates is None:
        noraf_json_rec.first('100').set('d', bibbi_person.dates)
        logger.info('[Noraf:%s] Setter 100$d = "%s"', noraf_person.id, bibbi_person.dates)

    elif bibbi_person.dates is not None and noraf_person.dates is not None:
        if re.sub('[^0-9]', '', bibbi_person.dates) != re.sub('[^0-9]', '', noraf_person.dates):
            logger.warning('Ulike datoer: "%s" (Noraf:%s) != "%s" (Bibbi:%s)',
                           noraf_person.dates,
                           noraf_person.id,
                           bibbi_person.dates,
                           bibbi_person.id)

    # 375 Gender
    if bibbi_person.gender is not None and noraf_person.gender is None:
        noraf_json_rec.add(NorafJsonMarcField.construct('375', ' ', ' ', [
            {
                "subcode": "a",
                "value": bibbi_person.gender,
            },
        ]))
        logger.info('[Noraf:%s] Setter 375$a = "%s"', noraf_person.id,
                    bibbi_person.gender)

    # ------------------------------------------------------------------------------------------------------------
    # 386 Nasjonalitet

    if bibbi_person.nationality is not None and noraf_person.nationality is None:
        noraf_json_rec.add(NorafJsonMarcField.construct('386', ' ', ' ', [
            {
                "subcode": "a",
                "value": bibbi_person.nationality,
            },
            {
                "subcode": "m",
                "value": "Nasjonalitet/regional gruppe",
            },
            {
                "subcode": "2",
                "value": "bibbi",
            },
        ]))
        logger.info('[Noraf:%s] Setter 386$a = "%s"', noraf_person.id,
                    bibbi_person.nationality)

    # ------------------------------------------------------------------------------------------------------------
    # 043 $c Landskoder

    if len(bibbi_person.country_codes) != 0 and len(noraf_person.country_codes) == 0:
        noraf_json_rec.add(NorafJsonMarcField.construct('043', ' ', ' ', [
            {
                "subcode": "c",
                "value": countryCode,
            }
            for countryCode in bibbi_person.country_codes
        ]))
        logger.info('[Noraf:%s] Setter 043$c = "%s"',
                    noraf_person.id, '" $c = "'.join(bibbi_person.country_codes))

    # ------------------------------------------------------------------------------------------------------------
    # Write result to file

    if dry_run:
        logger.debug(noraf_json_rec.as_dict())
    else:
        noraf.put(noraf_json_rec, reason='Oppdatering fra update_persons.py')
        if backup_path is not None:
            file_path: PathType = backup_path.joinpath('%s_after.json' % noraf_person.id)
            noraf_json_rec = noraf.get(match.noraf_id)
            with open(file_path, 'w', encoding='utf-8') as fp:
                fp.write(noraf_json_rec.as_json())

        time.sleep(5)
    return True


def main():
    """
    Scriptet oppdaterer personposter i Bibbi (via SQL) og Noraf (via REST-API) basert på inputt
    fra Excel-filen noraf_forslag_fra_isbn_tittel_match.xlsx
    """
    parser = argparse.ArgumentParser(description='Update persons using data from Excel sheet')
    parser.add_argument(
        'infile',
        nargs='?',
        default=r'/mnt/c/Users/dmheg/Bibliotekenes IT-Senter/Bibliotekfaglig avdeling - General/' +
                r'Promusdugnad/noraf_forslag_fra_isbn_tittel_match.xlsx',
        type=argparse.FileType('rb')
    )
    parser.add_argument('-n', '--dry', action='store_true', help='Dry run: Show the update operations the script ' +
                                                                 'would perform, without actually performing them.')
    parser.add_argument('-v', '--verbose', action='store_true', help='More verbose output.')
    args = parser.parse_args()

    if args.verbose:
        logger.setLevel(logging.DEBUG)
    else:
        logger.setLevel(logging.INFO)

    load_dotenv()
    promus = Promus()
    noraf = Noraf(os.getenv('BARE_KEY'))

    wb = load_workbook(args.infile)

    sheet: Worksheet = wb['Resultater']

    logger.info('Read %d rows, %d columns' % (sheet.max_row, sheet.max_column))
    marked_ok = []
    marked_not_ok = []
    for row in sheet.iter_rows(3):
        if row[0].value is None:
            continue

        status = str(row[0].value)

        if status.lower() == 'ok':
            marked_ok.append(BibbiNorafMatch(
                bibbi_id=str(row[1].value),
                noraf_id=str(row[6].value)
            ))
        else:
            logger.debug('Annen status: %s', status)
            marked_not_ok.append(BibbiNorafMatch(
                bibbi_id=str(row[1].value),
                noraf_id=str(row[6].value)
            ))

    logger.info('Records marked OK: %d, not OK: %d', len(marked_ok), len(marked_not_ok))

    backup_path = Path('./', 'backup')
    logger.info('Backup path: %s', backup_path.absolute())
    backup_path.mkdir(exist_ok=True)

    if args.dry:
        logger.info('Running in dry-run mode. No actual changes will be carried out.')
        time.sleep(2)
    else:
        if not questionary.confirm('Scriptet vil gjøre endringer i poster. Sikker på at du vil fortsette?', default=False).ask():
            return

    n = 0

    max_records = 100

    for match in marked_ok:
        if update_person(promus, noraf, match, args.dry, backup_path):
            n += 1
            if n >= max_records:
                logger.info('Processed %d records, will exit', n)
                return
