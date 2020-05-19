from __future__ import annotations

from dataclasses import asdict
from typing import Union
import sys
import time
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from requests import Session
from cachecontrol import CacheControlAdapter
from cachecontrol.heuristics import ExpiresAfter
from ..common.alma import get_alma_candidates
from ..common.interfaces import BarePerson, ViafPerson, Strategy
from ..matcher.matchers import Match, isbn_matcher, title_matcher, NoMatch
from ..common.viaf import get_viaf_candidates
from ..common.promus import Promus, BibbiPersons, BibbiPerson


def match_person(bibbi_person: BibbiPerson) -> Union[Match, NoMatch]:
    strategies = [
        Strategy(
            name='isbn',
            query='alma.isbn="{isbn}"',
            provider=get_alma_candidates,
            matcher=isbn_matcher,
        ),
        Strategy(
            name='creator+title',
            provider=get_alma_candidates,
            matcher=title_matcher,
            query='alma.creator="{creator}" AND alma.title="{title}"',
        ),
        Strategy(
            name='creator+fuzzy title',
            provider=get_alma_candidates,
            matcher=title_matcher,
            query='alma.creator="{creator}"',
        ),
        Strategy(
            name='viaf:creator+title',
            provider=get_viaf_candidates,
            matcher=title_matcher,
            query='local.personalNames="{creator}"',
        ),
    ]

    sys.stdout.write('%s %s' % (bibbi_person.name, bibbi_person.dates))

    viaf_match = None

    session = Session()
    session.mount('https://', CacheControlAdapter(
        heuristic=ExpiresAfter(days=1),
        max_retries=10
    ))

    for strategy in strategies:
        for bibbi_item in bibbi_person.items:

            sys.stdout.write('\n  [strategy:%s] %s "%s"' % (strategy.name, bibbi_item.isbn, '" || "'.join(bibbi_item.titles)))

            query = strategy.query.format(**{
                'creator': bibbi_person.name.strip(),
                'title': bibbi_item.titles[0].replace('"', '').strip(),
                **asdict(bibbi_item)
            })

            candidates = strategy.provider(query, session)

            for candidate in candidates:
                if match := strategy.matcher(bibbi_person, bibbi_item, candidate, strategy):
                    if isinstance(match.target, BarePerson):
                        sys.stdout.write(' -> MATCH!\n')
                        return match
                    elif isinstance(match.target, ViafPerson) and viaf_match is None:
                        # If we find a VIAF-only match, we keep it, but we will continue to check if we can
                        # find one with a BARE link.
                        viaf_match = match

    if viaf_match is not None:
        sys.stdout.write(' -> viaf only match\n')
        return viaf_match

    sys.stdout.write(' -> no match\n')
    return NoMatch()


def match_persons(persons: BibbiPersons):
    wb = Workbook()
    ws = wb.active

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 30
    ws.column_dimensions['K'].width = 30
    ws.column_dimensions['L'].width = 40
    ws.column_dimensions['M'].width = 40

    headers = [
        [
            'Bibbi-autoritet', '', '', '', '',
            'BARE-autoritetskandidat', '', '',
            'BARE-match basert på', '', '', '',
            'VIAF-autoritetskandidat', ''
        ],
        [
            'ID', 'Navn', 'Datoer', 'Antall poster', 'Nyeste post godkjent',
            'ID', 'Navn', 'Datoer', 'Strategi', 'ISBN/Tittel', 'Navn', 'Fødselsår',
            'ID', 'Tittel'
        ],
    ]

    header_font = Font(bold=True)

    for n1, v1 in enumerate(headers):
        for n2, v2 in enumerate(v1):
            cell = ws.cell(n1 + 1, n2 + 1, v2)
            cell.font = header_font
            cell.fill = PatternFill("solid", fgColor="FFFFEE")

    row = 3
    person_no = 0
    t0 = time.time()
    person_t = len(persons)
    for bibbi_id, bibbi_person in persons.items():
        person_no += 1

        dt = time.time() - t0
        time_s = datetime.now().strftime('%H:%M')
        print('[%s] %d/%d in %d secs' % (time_s, person_no, person_t, dt))

        match = match_person(bibbi_person)

        ws.cell(row=row, column=1, value=bibbi_id)
        ws.cell(row=row, column=2, value=bibbi_person.name)
        ws.cell(row=row, column=3, value=bibbi_person.dates)
        ws.cell(row=row, column=4, value=len(bibbi_person.items))
        ws.cell(row=row, column=5, value=bibbi_person.newest_approved)

        ws.cell(row=row, column=6, value=match.strategy)

        if isinstance(match.target, BarePerson):
            ws.cell(row=row, column=7, value=match.target.id)
            ws.cell(row=row, column=8, value=match.target.name)
            ws.cell(row=row, column=9, value=match.target.dates)
            ws.cell(row=row, column=10, value=match.title_similarity)
            ws.cell(row=row, column=11, value=match.name_similarity)
            ws.cell(row=row, column=12, value=match.date_similarity)

        if isinstance(match.target, ViafPerson):
            ws.cell(row=row, column=13, value='=HYPERLINK("%s")' % ('https://viaf.org/viaf/' + match.target.id))
            ws.cell(row=row, column=14, value=match.title_similarity)

        row += 1

        wb.save('results/bibbi-persons-match-alma.xlsx')


def main():
    load_dotenv()

    promus = Promus()

    # Hent alle personer fra Bibbi
    bibbi_persons = promus.fetch_persons()

    # Velg ut de som har minst en utgivelse i 2019 eller 2020, men spar på alle utgivelsene til disse personene,
    # så vi kan bruke dem til matching.
    bibbi_persons_filtered = {}
    for bibbi_id, bibbi_person in bibbi_persons.items():
        newest_year = int(bibbi_person.newest_approved.strftime('%Y'))
        if 2019 <= newest_year <= 2020:
            bibbi_persons_filtered[bibbi_id] = bibbi_person

    print('Fetched %d persons from Promus' % len(bibbi_persons))
    print('of which %d persons having items published within requested date range' % len(bibbi_persons_filtered))

    match_persons(bibbi_persons_filtered)
